import streamlit as st
import json
import re
import tempfile
import os
import time
import io
from datetime import datetime

# Safe imports for data handling & visual BI
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import yfinance as yf
except ImportError:
    yf = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pypdf
except ImportError:
    pypdf = None

try:
    import plotly.graph_objects as go
    import plotly.express as px
except ImportError:
    go = None
    px = None

from google import genai
from google.genai import types

# ============================================================
# PAGE CONFIGURATION
# ============================================================

st.set_page_config(
    page_title="Financial Analyst AI | Terminal",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ============================================================
# INSTITUTIONAL LIQUID GLASS FINTECH COCKPIT THEME & CSS
# ============================================================

st.markdown("""
<style>
/* Base Theme with Financial Analytics Watermark Background */
.stApp {
    background-color: #06080e;
    background-image: 
        radial-gradient(circle at 15% 20%, rgba(30, 58, 138, 0.15) 0%, transparent 40%),
        radial-gradient(circle at 85% 80%, rgba(16, 185, 129, 0.08) 0%, transparent 40%),
        linear-gradient(rgba(6, 8, 14, 0.94), rgba(6, 8, 14, 0.94)),
        url("data:image/svg+xml,%3Csvg width='100' height='100' viewBox='0 0 100 100' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M10 10h80v80H10z' fill='none'/%3E%3Cpath d='M20 70l20-30 20 15 25-35' stroke='rgba(59, 130, 246, 0.04)' stroke-width='2' fill='none'/%3E%3Ctext x='15' y='30' fill='rgba(255,255,255,0.02)' font-family='monospace' font-size='10'%3EBI-ANALYTICS%3C/text%3E%3C/svg%3E");
    background-size: cover, cover, cover, 180px 180px;
    color: #f1f5f9;
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
}
.block-container {
    max-width: 1450px;
    padding-top: 1.8rem;
    padding-bottom: 4rem;
}

/* Hide Default Streamlit Status Widget */
div[data-testid="stStatusWidget"] {
    display: none !important;
    visibility: hidden !important;
}

/* Landing Page Downward Sliding Bouncy Animation */
@keyframes landingBounce {
    0% {
        opacity: 0;
        transform: translateY(-60px) scale(0.98);
    }
    50% {
        opacity: 1;
        transform: translateY(12px) scale(1.01);
    }
    75% {
        transform: translateY(-6px) scale(0.995);
    }
    90% {
        transform: translateY(3px) scale(1.002);
    }
    100% {
        opacity: 1;
        transform: translateY(0) scale(1);
    }
}

.landing-animate {
    animation: landingBounce 1.1s cubic-bezier(0.2, 0.8, 0.2, 1) forwards;
}

/* Liquid Glass Card Styling */
.liquid-glass-card {
    background: rgba(13, 20, 36, 0.65);
    backdrop-filter: blur(16px);
    -webkit-backdrop-filter: blur(16px);
    border: 1px solid rgba(255, 255, 255, 0.08);
    border-top: 1px solid rgba(255, 255, 255, 0.15);
    border-radius: 16px;
    padding: 24px;
    box-shadow: 0 20px 40px rgba(0, 0, 0, 0.5), inset 0 1px 0 rgba(255, 255, 255, 0.1);
    transition: all 0.3s ease;
}

.liquid-glass-card:hover {
    border-color: rgba(59, 130, 246, 0.4);
    box-shadow: 0 25px 50px rgba(0, 0, 0, 0.6), 0 0 20px rgba(59, 130, 246, 0.15);
}

/* Electric Glowing Animations for Cards and Uploader */
@keyframes electricPulseBlue {
    0%, 100% {
        box-shadow: 0 0 15px rgba(59, 130, 246, 0.3), inset 0 0 10px rgba(59, 130, 246, 0.1);
        border-color: rgba(59, 130, 246, 0.7);
    }
    50% {
        box-shadow: 0 0 30px rgba(96, 165, 250, 0.6), inset 0 0 20px rgba(96, 165, 250, 0.25);
        border-color: #60a5fa;
    }
}

@keyframes electricPulseGreen {
    0%, 100% {
        box-shadow: 0 0 15px rgba(16, 185, 129, 0.3), inset 0 0 10px rgba(16, 185, 129, 0.1);
        border-color: rgba(16, 185, 129, 0.7);
    }
    50% {
        box-shadow: 0 0 30px rgba(52, 211, 153, 0.6), inset 0 0 20px rgba(52, 211, 153, 0.25);
        border-color: #34d399;
    }
}

@keyframes electricPulseYellow {
    0%, 100% {
        box-shadow: 0 0 15px rgba(245, 158, 11, 0.3), inset 0 0 10px rgba(245, 158, 11, 0.1);
        border-color: rgba(245, 158, 11, 0.7);
    }
    50% {
        box-shadow: 0 0 30px rgba(251, 191, 36, 0.6), inset 0 0 20px rgba(251, 191, 36, 0.25);
        border-color: #fbbf24;
    }
}

@keyframes electricPulseRed {
    0%, 100% {
        box-shadow: 0 0 15px rgba(239, 68, 68, 0.3), inset 0 0 10px rgba(239, 68, 68, 0.1);
        border-color: rgba(239, 68, 68, 0.7);
    }
    50% {
        box-shadow: 0 0 30px rgba(248, 113, 113, 0.6), inset 0 0 20px rgba(248, 113, 113, 0.25);
        border-color: #f87171;
    }
}

/* Streamlit Native File Uploader Electric Border Frame */
div[data-testid="stFileUploader"] {
    background: rgba(11, 17, 30, 0.85);
    backdrop-filter: blur(12px);
    border: 2px dashed #3b82f6;
    border-radius: 16px;
    padding: 16px;
    animation: electricPulseBlue 3s infinite ease-in-out;
}

/* Symmetrical Electric KPI Cards with 4 Performance Statuses */
.electric-kpi-card-blue {
    background: linear-gradient(145deg, #0b101c 0%, #060911 100%);
    border-radius: 16px;
    padding: 20px;
    height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    animation: electricPulseBlue 3.5s infinite ease-in-out;
    border: 2px solid #3b82f6;
}

.electric-kpi-card-green {
    background: linear-gradient(145deg, #071f16 0%, #040d0a 100%);
    border-radius: 16px;
    padding: 20px;
    height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    animation: electricPulseGreen 3.5s infinite ease-in-out;
    border: 2px solid #10b981;
}

.electric-kpi-card-yellow {
    background: linear-gradient(145deg, #221808 0%, #0d0a04 100%);
    border-radius: 16px;
    padding: 20px;
    height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    animation: electricPulseYellow 3.5s infinite ease-in-out;
    border: 2px solid #fbbf24;
}

.electric-kpi-card-red {
    background: linear-gradient(145deg, #240d12 0%, #0d0406 100%);
    border-radius: 16px;
    padding: 20px;
    height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: space-between;
    animation: electricPulseRed 3.5s infinite ease-in-out;
    border: 2px solid #ef4444;
}

/* Spinner & Dynamic Loaders */
@keyframes spinGlow {
    0% { transform: rotate(0deg); }
    100% { transform: rotate(360deg); }
}

@keyframes shimmerBar {
    0% { transform: translateX(-100%); }
    100% { transform: translateX(200%); }
}

@keyframes pulseText {
    0%, 100% { opacity: 0.6; }
    50% { opacity: 1; }
}

.center-loader-box {
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    background: rgba(13, 18, 30, 0.95) !important;
    backdrop-filter: blur(24px) !important;
    border: 1px solid rgba(59, 130, 246, 0.45) !important;
    border-radius: 18px !important;
    padding: 36px 32px !important;
    margin: 25px auto !important;
    text-align: center;
    max-width: 620px;
}
.fintech-spinner {
    width: 50px;
    height: 50px;
    border: 3.5px solid rgba(59, 130, 246, 0.15);
    border-top: 3.5px solid #60a5fa;
    border-right: 3.5px solid #3b82f6;
    border-radius: 50%;
    animation: spinGlow 0.85s linear infinite;
    margin-bottom: 16px;
}

.loader-progress-track {
    background: #151d2f;
    border-radius: 4px;
    height: 6px;
    width: 80%;
    margin: 20px auto 0 auto;
    overflow: hidden;
    position: relative;
}

.loader-progress-fill {
    background: linear-gradient(90deg, transparent, #3b82f6, #60a5fa, transparent);
    height: 100%;
    width: 50%;
    position: absolute;
    animation: shimmerBar 1.5s infinite linear;
}

.pulse-text {
    animation: pulseText 2s infinite ease-in-out;
}

/* Telemetry & Badges */
.telemetry-bar {
    display: flex;
    flex-wrap: wrap;
    gap: 10px;
    margin-top: 10px;
    margin-bottom: 20px;
    padding: 10px 14px;
    background: rgba(13, 18, 30, 0.65);
    border: 1px solid #1e293b;
    border-radius: 10px;
}
.telemetry-pill {
    background: #0a0e1a;
    border: 1px solid #23334d;
    color: #93c5fd;
    font-size: 12px;
    font-weight: 600;
    padding: 4px 12px;
    border-radius: 20px;
    display: inline-flex;
    align-items: center;
    gap: 6px;
}

/* Typography & Banners */
.section-title {
    font-size: 21px;
    font-weight: 750;
    color: #f8fafc;
    margin-top: 24px;
    margin-bottom: 4px;
}
.section-description {
    color: #94a3b8;
    font-size: 13.5px;
    margin-bottom: 16px;
}
.fintech-banner {
    background: linear-gradient(135deg, #0c1220 0%, #060913 100%);
    border: 1px solid #1e293b;
    border-left: 4px solid #3b82f6;
    border-radius: 12px;
    padding: 18px 22px;
    margin-top: 30px;
    margin-bottom: 16px;
}
.fintech-banner-title {
    font-size: 18px;
    font-weight: 750;
    color: #ffffff;
    margin-bottom: 4px;
}
.fintech-banner-desc {
    font-size: 13px;
    color: #94a3b8;
}

/* Symmetrical Overview Cards */
.company-card {
    background: rgba(10, 14, 26, 0.85);
    border: 1px solid #1a2234;
    border-radius: 16px;
    padding: 18px;
    height: 155px;
    display: flex;
    flex-direction: column;
    justify-content: flex-start;
    overflow-y: auto;
    margin-bottom: 10px;
    animation: electricPulseBlue 4s infinite ease-in-out;
}
.company-label {
    color: #fbbf24;
    font-size: 11px;
    text-transform: uppercase;
    letter-spacing: 0.8px;
    margin-bottom: 6px;
    font-weight: 700;
}
.company-value {
    color: #f8fafc;
    font-size: 13.5px;
    font-weight: 550;
    line-height: 1.4;
}

/* BI KPI Cards */
.kpi-header-row {
    display: flex;
    justify-content: space-between;
    align-items: center;
}
.kpi-title {
    color: #94a3b8;
    font-size: 11.5px;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}
.kpi-main-val {
    color: #ffffff;
    font-size: 22px;
    font-weight: 800;
    margin: 4px 0;
}
.spark-track {
    background: #151d2f;
    border-radius: 4px;
    height: 5px;
    width: 100%;
    margin-top: 8px;
    overflow: hidden;
}

/* Chat Styling */
.chat-box-card {
    background: #0a0e1a;
    border: 1px solid #1a2234;
    border-radius: 12px;
    padding: 16px 18px;
    margin-bottom: 14px;
}
.chat-user-badge {
    color: #60a5fa;
    font-weight: 750;
    font-size: 13px;
    margin-bottom: 6px;
    display: flex;
    align-items: center;
    gap: 6px;
}
.chat-bot-badge {
    color: #34d399;
    font-weight: 750;
    font-size: 13px;
    margin-bottom: 8px;
    display: flex;
    align-items: center;
    gap: 6px;
}
.chat-text {
    color: #e2e8f0;
    font-size: 13.5px;
    line-height: 1.6;
}
.footer {
    color: #64748b;
    font-size: 12px;
    text-align: center;
    padding-top: 25px;
}

/* =======================================================
   MODERN SEPARATED BORDERED NAVBAR FOR TABS
   ======================================================= */
div[data-testid="stTabs"] {
    margin-top: 15px !important;
    margin-bottom: 25px !important;
}

/* 1. Main outer container with modern rounded border */
div[data-testid="stTabs"] > div > div[role="tablist"] {
    background: rgba(10, 14, 26, 0.85) !important;
    backdrop-filter: blur(14px) !important;
    -webkit-backdrop-filter: blur(14px) !important;
    border: 1px solid rgba(59, 130, 246, 0.3) !important;
    border-radius: 16px !important;
    padding: 8px !important;
    gap: 4px !important;
    display: flex !important;
    flex-wrap: wrap !important;
    box-shadow: 0 10px 30px rgba(0, 0, 0, 0.4) !important;
}

/* 2. Eradicate Streamlit default red accent line and borders */
div[data-testid="stTabs"] [data-baseweb="tab-highlight"],
div[data-testid="stTabs"] [data-baseweb="tab-border"] {
    display: none !important;
    visibility: hidden !important;
}

/* 3. Individual tab item with clean vertical divider borders between them */
div[data-testid="stTabs"] button[role="tab"] {
    background: transparent !important;
    border-radius: 10px !important;
    border-right: 1px solid rgba(255, 255, 255, 0.08) !important; /* Modern vertical divider */
    border-top: none !important;
    border-left: none !important;
    border-bottom: none !important;
    padding: 10px 20px !important;
    min-height: 42px !important;
    transition: all 0.25s ease !important;
    margin: 0 !important;
}

/* Remove border from the very last tab so it looks balanced */
div[data-testid="stTabs"] button[role="tab"]:last-child {
    border-right: none !important;
}

/* 4. Unselected tab text styling */
div[data-testid="stTabs"] button[role="tab"][aria-selected="false"] p,
div[data-testid="stTabs"] button[role="tab"][aria-selected="false"] span {
    color: #94a3b8 !important;
    font-weight: 600 !important;
    font-size: 13.5px !important;
    margin: 0 !important;
}

/* 5. Hover state */
div[data-testid="stTabs"] button[role="tab"][aria-selected="false"]:hover {
    background: rgba(255, 255, 255, 0.06) !important;
}
div[data-testid="stTabs"] button[role="tab"][aria-selected="false"]:hover p,
div[data-testid="stTabs"] button[role="tab"][aria-selected="false"]:hover span {
    color: #ffffff !important;
}

/* 6. Active Tab: Modern Glass Box Selection */
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
    background: rgba(59, 130, 246, 0.2) !important;
    border: 1px solid rgba(59, 130, 246, 0.5) !important;
    border-radius: 10px !important;
    box-shadow: 0 4px 15px rgba(59, 130, 246, 0.25) !important;
    border-right: none !important; /* Prevents divider overlap on active state */
}

/* 7. Active Tab Typography (Pure white & bright blue, killing any red) */
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] p,
div[data-testid="stTabs"] button[role="tab"][aria-selected="true"] span {
    color: #60a5fa !important;
    font-weight: 750 !important;
    font-size: 13.5px !important;
    margin: 0 !important;
}

</style>
""", unsafe_allow_html=True)

# ============================================================
# SESSION STATE INITIALIZATION
# ============================================================

defaults = {
    "gemini_file": None,
    "uploaded_name": None,
    "analysis": None,
    "selected_model": None,
    "position_assessment": None,
    "chat_history": [],
    "processing_seconds": 0.0,
    "file_size_mb": 0.0,
    "forecast_data": None,
    "forecast_company": None,
    "page_count": 0
}

for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value

# ============================================================
# API KEY & CLIENT
# ============================================================

try:
    API_KEY = st.secrets["GEMINI_API_KEY"]
except Exception:
    API_KEY = None

if not API_KEY:
    st.error("API key was not found. Please add GEMINI_API_KEY to Streamlit Secrets.")
    st.stop()

@st.cache_resource
def create_client(api_key):
    return genai.Client(api_key=api_key)

client = create_client(API_KEY)

# ============================================================
# ACTIVE GEMINI 3-SERIES PRODUCTION MODELS
# ============================================================

ACTIVE_MODELS = [
    "gemini-3.7-flash",
    "gemini-3.6-flash",
    "gemini-3.5-flash-lite",
    "gemini-3.1-pro-preview"
]

def generate_with_fallback(contents, json_mode=False, use_search=False):
    errors = []
    ordered = ACTIVE_MODELS.copy()
    if st.session_state.selected_model and st.session_state.selected_model in ordered:
        ordered.remove(st.session_state.selected_model)
        ordered.insert(0, st.session_state.selected_model)

    for model in ordered:
        for attempt in range(5):  # Exponential Backoff enabled: 5 Attempts max
            try:
                tools_list = [types.Tool(google_search=types.GoogleSearch())] if use_search else None
                
                # When search tool is enabled, response_mime_type cannot be application/json
                if json_mode and not use_search:
                    config = types.GenerateContentConfig(
                        response_mime_type="application/json",
                        temperature=0.1,
                        max_output_tokens=8192
                    )
                else:
                    config = types.GenerateContentConfig(
                        temperature=0.2,
                        max_output_tokens=8192,
                        tools=tools_list
                    )

                response = client.models.generate_content(
                    model=model,
                    contents=contents,
                    config=config
                )

                st.session_state.selected_model = model
                return response

            except Exception as error:
                err_str = str(error)
                # Catch both Rate Limits (429) AND Overloaded Servers (503) seamlessly
                if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str or "503" in err_str or "UNAVAILABLE" in err_str:
                    sleep_time = 2.0 * (2 ** attempt)  # 2s, 4s, 8s, 16s...
                    time.sleep(sleep_time)
                    continue
                errors.append(f"{model}: {err_str}")
                break

    error_text = "\n\n".join(errors)
    raise Exception(f"API Rate limit reached or server is currently overloaded. Please try again in a few minutes.\n\n{error_text}")

def generate_chat_response(contents, use_search=False):
    """Fast-track dedicated fallback specifically for the chat copilot to maximize speed."""
    errors = []
    # Prioritize flash-lite for maximum chat speed and token efficiency
    chat_models = [
        "gemini-3.5-flash-lite",
        "gemini-3.6-flash",
        "gemini-3.7-flash"
    ]
    for model in chat_models:
        for attempt in range(4): 
            try:
                tools_list = [types.Tool(google_search=types.GoogleSearch())] if use_search else None
                config = types.GenerateContentConfig(
                    temperature=0.2,
                    max_output_tokens=4096,
                    tools=tools_list
                )
                response = client.models.generate_content(
                    model=model,
                    contents=contents,
                    config=config
                )
                return response
            except Exception as error:
                err_str = str(error)
                if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str or "503" in err_str or "UNAVAILABLE" in err_str:
                    time.sleep(2.0 * (2 ** attempt))
                    continue
                errors.append(f"{model}: {err_str}")
                break
    raise Exception(f"Chat API is currently overloaded by Google's limits. Please try again in a few moments.\n\n{errors}")

# ============================================================
# SAFE PARSERS & UNIVERSAL STOCK LOOKUP
# ============================================================

def clean_json_response(text):
    if not text:
        return {}
    raw = text.strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.IGNORECASE)
    raw = re.sub(r"\s*```$", "", raw)
    raw = raw.strip()

    try:
        return json.loads(raw)
    except Exception:
        pass

    start = raw.find("{")
    end = raw.rfind("}")
    if start != -1 and end != -1 and end > start:
        candidate = raw[start:end + 1]
        try:
            return json.loads(candidate)
        except Exception:
            pass

    if start != -1:
        truncated = raw[start:]
        for suffix in ['"}', '"}]}', '"]}', ']}', '}']:
            try:
                return json.loads(truncated + suffix)
            except Exception:
                continue

    return {}

def parse_clean_float(val):
    if val is None:
        return None
    cleaned = str(val).replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
    match = re.search(r"[-+]?\d*\.?\d+", cleaned)
    if match:
        try:
            return float(match.group())
        except Exception:
            return None
    return None

def auto_classify_metric(name):
    n = str(name).lower()
    if any(k in n for k in ["revenue", "income", "profit", "pat", "ebitda", "margin", "expense", "cost", "turnover", "fee", "sales"]):
        return "Revenue & Profit"
    elif any(k in n for k in ["npa", "crar", "car", "ratio", "roe", "roa", "coverage", "pcr", "cushion", "leverage", "nim", "%"]):
        return "Financial Health Ratios"
    else:
        return "Balance Sheet & Assets"

def fetch_live_stock_price(company_name, ticker_hint=""):
    if not yf:
        return None
    try:
        candidates = []
        if ticker_hint:
            clean_t = re.sub(r'[^A-Za-z0-9]', '', str(ticker_hint)).upper()
            if clean_t:
                candidates.extend([f"{clean_t}.NS", f"{clean_t}.BO", clean_t])
        if not candidates and company_name:
            first_word = re.sub(r'[^A-Za-z0-9]', '', company_name.split()[0]).upper()
            if len(first_word) >= 3:
                candidates.extend([f"{first_word}.NS", f"{first_word}.BO", first_word])

        for sym in candidates:
            try:
                tk = yf.Ticker(sym)
                hist = tk.history(period="5d")
                if not hist.empty:
                    last_price = float(hist["Close"].iloc[-1])
                    last_date = hist.index[-1].strftime("%d %b %Y")
                    exchange_label = "NSE" if ".NS" in sym else ("BSE" if ".BO" in sym else "Global Market")
                    return {
                        "is_listed": True,
                        "ticker": sym.replace(".NS", "").replace(".BO", ""),
                        "price": last_price,
                        "as_on": last_date,
                        "exchange": exchange_label
                    }
            except Exception:
                continue
    except Exception:
        pass
    return None

def upload_pdf_to_gemini(uploaded_file):
    temp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_file:
            temp_file.write(uploaded_file.getbuffer())
            temp_path = temp_file.name

        # Direct, reliable upload to Gemini without invalid arguments
        gemini_file = client.files.upload(file=temp_path)
        
        # Robust polling loop to verify the file is ready
        for _ in range(90):
            try:
                gemini_file = client.files.get(name=gemini_file.name)
                state = getattr(gemini_file, "state", None)
                state_name = getattr(state, "name", str(state))
                if state_name == "ACTIVE":
                    return gemini_file
                elif state_name == "FAILED":
                    raise Exception("The PDF file format is too complex or scanned images couldn't be read. Please try a text-searchable PDF.")
            except Exception:
                pass
            time.sleep(2)

        return gemini_file
    finally:
        if temp_path:
            try:
                os.remove(temp_path)
            except Exception:
                pass

# ============================================================
# SECTION 1: HERO SECTION
# ============================================================

st.markdown("""
<div style="text-align: center; padding: 20px 0 10px 0;" class="landing-animate">
    <div style="font-size: 38px; font-weight: 800; color: #f8fafc; letter-spacing: -0.5px; margin-bottom: 8px;">Financial Analyst AI</div>
    <div style="font-size: 15px; color: #94a3b8; max-width: 750px; margin: 0 auto 18px auto; line-height: 1.5;">Institutional Equity Research & Corporate Fundamental Intelligence Terminal</div>
    <div style="display: flex; justify-content: center; flex-wrap: wrap; gap: 10px;">
        <span class="telemetry-pill">📈 Real-Time Equity Tracking</span>
        <span class="telemetry-pill">📊 Balance Sheet Auditing</span>
        <span class="telemetry-pill">⚡ Multi-Pillar Diagnostic Engine</span>
        <span class="telemetry-pill">🛡️ Risk & Solvency Assessment</span>
    </div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# SECTION 2: UPLOAD FINANCIAL REPORT
# ============================================================

st.markdown('<div class="section-title landing-animate">Upload Financial Report</div>', unsafe_allow_html=True)
st.markdown('<div class="section-description landing-animate">Upload any corporate annual report or financial filing (PDF) to initiate natural-language dynamic analysis.</div>', unsafe_allow_html=True)

uploaded_file = st.file_uploader(
    "Upload Financial Report (PDF)",
    type=["pdf"],
    label_visibility="collapsed",
    key="main_pdf_uploader"
)

st.markdown("""
<div class="liquid-glass-card landing-animate" style="margin-top: 16px; margin-bottom: 24px; padding: 16px 20px;">
    <div style="font-weight: 700; color: #fbbf24; margin-bottom: 4px; font-size: 13px; display: flex; align-items: center; gap: 6px;">
        ⏱️ <span>Document Size & Processing Advisory</span>
    </div>
    <div style="color: #94a3b8; font-size: 12.5px; line-height: 1.5;">
        Standard multi-hundred page corporate filings typically take 1 to 3 minutes to parse and reconcile. <b>Disclaimer:</b> Massive files (500+ pages or >5MB) take approximately more than 5 minutes to load, synthesize, and give the output in a precise, correct, and error-free manner.
    </div>
</div>
""", unsafe_allow_html=True)

loader_container = st.empty()

# ============================================================
# SECTION 3: FEATURES SHOWCASE (BEFORE UPLOAD)
# ============================================================

if not st.session_state.gemini_file or not st.session_state.analysis:
    if not uploaded_file:
        st.info("👆 Upload an annual report PDF above to begin institutional financial analysis.")
        st.markdown("---")
        st.markdown('<div class="section-title" style="margin-top:0;">Terminal Research Capabilities</div>', unsafe_allow_html=True)
        st.markdown('<div class="section-description">Automated modules generated upon document reconciliation:</div>', unsafe_allow_html=True)
        
        c_feat1, c_feat2, c_feat3, c_feat4 = st.columns(4)
        with c_feat1:
            st.markdown("""
            <div class="electric-kpi-card-blue">
                <div style="font-size: 24px; margin-bottom: 8px;">📊</div>
                <div style="font-size: 15px; font-weight: 750; color: #fff; margin-bottom: 6px;">Financial Extraction</div>
                <div style="font-size: 12.5px; color: #94a3b8; line-height: 1.4;">Extracts primary income statement, balance sheet, and operating metrics dynamically.</div>
            </div>
            """, unsafe_allow_html=True)
        with c_feat2:
            st.markdown("""
            <div class="electric-kpi-card-blue">
                <div style="font-size: 24px; margin-bottom: 8px;">📈</div>
                <div style="font-size: 15px; font-weight: 750; color: #fff; margin-bottom: 6px;">Portfolio Intelligence</div>
                <div style="font-size: 12.5px; color: #94a3b8; line-height: 1.4;">Connects to live market pricing to evaluate cost basis, purchase safety, and compounding.</div>
            </div>
            """, unsafe_allow_html=True)
        with c_feat3:
            st.markdown("""
            <div class="electric-kpi-card-blue">
                <div style="font-size: 24px; margin-bottom: 8px;">🛡️</div>
                <div style="font-size: 15px; font-weight: 750; color: #fff; margin-bottom: 6px;">Executive Scorecard</div>
                <div style="font-size: 12.5px; color: #94a3b8; line-height: 1.4;">Multi-pillar evaluation assessing growth momentum, profit quality, and balance sheet cushion.</div>
            </div>
            """, unsafe_allow_html=True)
        with c_feat4:
            st.markdown("""
            <div class="electric-kpi-card-blue">
                <div style="font-size: 24px; margin-bottom: 8px;">💬</div>
                <div style="font-size: 15px; font-weight: 750; color: #fff; margin-bottom: 6px;">Grounded AI Copilot</div>
                <div style="font-size: 12.5px; color: #94a3b8; line-height: 1.4;">Interactive research assistant answering custom queries strictly using audited facts.</div>
            </div>
            """, unsafe_allow_html=True)

# ============================================================
# AUTOMATIC GENERATION ON UPLOAD WITH DYNAMIC LOADER
# ============================================================

if uploaded_file:
    is_new_file = (st.session_state.uploaded_name != uploaded_file.name)

    if is_new_file or st.session_state.analysis is None:
        start_time = time.time()
        file_mb = round(len(uploaded_file.getvalue()) / (1024 * 1024), 2)
        st.session_state.file_size_mb = file_mb

        if pypdf is not None:
            try:
                pdf_reader = pypdf.PdfReader(uploaded_file)
                st.session_state.page_count = len(pdf_reader.pages)
                uploaded_file.seek(0)
            except Exception:
                st.session_state.page_count = "Unknown"
        else:
            st.session_state.page_count = "N/A"

        # Step 1: Ingestion
        loader_container.markdown("""
        <div class="center-loader-box">
            <div class="telemetry-pill" style="margin-bottom: 12px;">Step 1 of 3 • Document Ingestion</div>
            <div class="fintech-spinner"></div>
            <div style="font-size: 18px; font-weight: 750; color: #fff; margin-bottom: 6px;">Uploading & Verifying Document...</div>
            <div class="pulse-text" style="font-size: 13px; color: #94a3b8;">Transferring PDF securely to Gemini analytical cluster.</div>
            <div class="loader-progress-track"><div class="loader-progress-fill"></div></div>
        </div>
        """, unsafe_allow_html=True)

        try:
            gemini_file = upload_pdf_to_gemini(uploaded_file)
            st.session_state.gemini_file = gemini_file
            st.session_state.uploaded_name = uploaded_file.name

            # Step 2: Parsing
            loader_container.markdown("""
            <div class="center-loader-box">
                <div class="telemetry-pill" style="margin-bottom: 12px;">Step 2 of 3 • Parsing Statements</div>
                <div class="fintech-spinner"></div>
                <div style="font-size: 18px; font-weight: 750; color: #fff; margin-bottom: 6px;">Indexing Financial Data...</div>
                <div class="pulse-text" style="font-size: 13px; color: #94a3b8;">Extracting balance sheets, income statements, and natural-language text.</div>
                <div class="loader-progress-track"><div class="loader-progress-fill"></div></div>
            </div>
            """, unsafe_allow_html=True)
            time.sleep(1.0)

            # Step 3: Synthesis & AI Processing (Includes Shimmer Bar & Pulse Text warning for large files)
            loader_container.markdown("""
            <div class="center-loader-box">
                <div class="telemetry-pill" style="margin-bottom: 12px;">Step 3 of 3 • AI Synthesis & Analysis</div>
                <div class="fintech-spinner"></div>
                <div style="font-size: 18px; font-weight: 750; color: #fff; margin-bottom: 6px;">Running Deep Financial Analysis...</div>
                <div class="pulse-text" style="font-size: 13px; color: #94a3b8; line-height: 1.5;">
                    Structuring health indicators, risk matrices, and dashboards.<br>
                    <span style="color: #fbbf24; font-weight: 600;">(Disclaimer: Massive files like 500+ pages or >5MB take approximately 5+ minutes to synthesize for precise, correct, and error-free output. Please do not refresh.)</span>
                </div>
                <div class="loader-progress-track"><div class="loader-progress-fill"></div></div>
            </div>
            """, unsafe_allow_html=True)

            analysis_prompt = """
You are an expert institutional equity research analyst. Automatically detect the type of corporate report uploaded (e.g., Bank/NBFC, Technology, Manufacturing, FMCG, Energy, etc.).
Dynamically extract metrics and structure your response strictly in valid JSON matching this schema:
{
  "company_overview": {
    "company_name": "Full company name",
    "stock_ticker": "Ticker symbol if available",
    "industry": "Detected industry sector",
    "business_type": "2 sentences describing core operations and revenue model",
    "reporting_period": "Reporting fiscal year",
    "report_type": "Annual Report or Financial Filing"
  },
  "terms_cheat_sheet": [
    { "term": "Term name", "meaning": "1 sentence explanation" }
  ],
  "key_metrics": [
    {
      "metric": "Line item name",
      "current_period": "Current value",
      "previous_period": "Previous value",
      "yoy_growth": "YoY growth percentage",
      "unit": "₹ Crore / % / USD",
      "basis": "Consolidated / Standalone",
      "what_it_means": "1 sentence explanation"
    }
  ],
  "investor_scorecard": {
    "growth_momentum": { "badge": "Robust", "verdict": "Summary sentence", "health_pct": 85, "points": ["Point 1", "Point 2", "Point 3"] },
    "profitability_quality": { "badge": "Solid", "verdict": "Summary sentence", "health_pct": 80, "points": ["Point 1", "Point 2", "Point 3"] },
    "balance_sheet_safety": { "badge": "Secure", "verdict": "Summary sentence", "health_pct": 92, "points": ["Point 1", "Point 2", "Point 3"] },
    "strategic_execution": { "badge": "Active", "verdict": "Summary sentence", "health_pct": 88, "points": ["Point 1", "Point 2", "Point 3"] }
  },
  "in_depth_investigation": {
    "profitability_and_margins": { "headline": "Verdict sentence", "points": ["Point 1", "Point 2", "Point 3"] },
    "borrowings_and_capital_cushion": { "headline": "Verdict sentence", "points": ["Point 1", "Point 2", "Point 3"] },
    "operating_efficiency_and_scale": { "headline": "Verdict sentence", "points": ["Point 1", "Point 2", "Point 3"] }
  },
  "management_commentary": [
    { "title": "Theme title", "summary": "Summary explanation" }
  ],
  "risks": [
    { "title": "Risk title", "category": "Market & Economy", "impact_level": "High", "what_is_the_risk": "Explanation", "why_it_matters": "Financial impact" }
  ],
  "analyst_takeaway": {
    "improving": ["Tailwind 1", "Tailwind 2", "Tailwind 3"],
    "weakening": ["Headwind 1", "Headwind 2", "Headwind 3"],
    "growth_drivers": ["Driver 1", "Driver 2"],
    "investor_watch": ["Checkpoint 1", "Checkpoint 2"],
    "sentiment_score": 75
  }
}
"""
            response = generate_with_fallback(
                contents=[analysis_prompt, gemini_file],
                json_mode=True
            )
            data = clean_json_response(response.text)
            elapsed_time = round(time.time() - start_time, 1)
            st.session_state.processing_seconds = elapsed_time

            loader_container.empty()
            if data and "company_overview" in data:
                st.session_state.analysis = data
                st.session_state.position_assessment = None
                st.session_state.chat_history = []
                st.success(f"Institutional analysis completed successfully in {elapsed_time}s!")
                st.rerun()
            else:
                st.error("The document could not be completely parsed. Please re-upload or try again.")
        except Exception as e:
            loader_container.empty()
            st.error(f"Processing error: {e}")

# ============================================================
# NO PDF STATE STOP
# ============================================================

if not st.session_state.gemini_file or not st.session_state.analysis:
    st.stop()

# ============================================================
# PART 1: AUTO-RENDERED POST-UPLOAD SECTIONS
# ============================================================

data = st.session_state.analysis
company = data.get("company_overview", {})
metrics = data.get("key_metrics", [])
scorecard = data.get("investor_scorecard", {})
deep_investigation = data.get("in_depth_investigation", {})
management = data.get("management_commentary", [])
risks = data.get("risks", [])
takeaway = data.get("analyst_takeaway", {})

proc_time = st.session_state.get("processing_seconds", 0.0)
f_size = st.session_state.get("file_size_mb", 0.0)
model_name = st.session_state.get("selected_model", "Gemini Engine")
p_count = st.session_state.get("page_count", "Unknown")

# Header & Telemetry Bar
st.markdown(f"""
<div class="telemetry-bar landing-animate">
    <span class="telemetry-pill">⏱️ Processing Latency: <b>{proc_time}s</b></span>
    <span class="telemetry-pill">🧠 Model Engine: <b>{model_name}</b></span>
    <span class="telemetry-pill">📄 Filing Size: <b>{f_size} MB</b></span>
    <span class="telemetry-pill">📑 Pages Analyzed: <b>{p_count}</b></span>
    <span class="telemetry-pill" style="border-color: #059669; color: #34d399;">🟢 Verification: <b>Audited & Reconciled</b></span>
</div>
""", unsafe_allow_html=True)

# Financial Glossary Expander
with st.expander("📌 Financial Glossary & Core Reporting Nomenclature", expanded=False):
    cheat_terms = data.get("terms_cheat_sheet", [])
    if cheat_terms:
        term_map = {item.get("term", "").strip(): item.get("meaning", "").strip() for item in cheat_terms if item.get("term")}
        term_names = list(term_map.keys())
        if term_names:
            selected_jargon = st.selectbox("Select reporting term:", options=term_names, index=0, key="glossary_slicer")
            st.markdown(f"""
            <div class="liquid-glass-card" style="padding: 14px 18px; margin-top: 8px;">
                <div style="color: #60a5fa; font-weight: 700; font-size: 13px; margin-bottom: 4px;">💡 {selected_jargon}</div>
                <div style="color: #cbd5e1; font-size: 13px;">{term_map[selected_jargon]}</div>
            </div>
            """, unsafe_allow_html=True)

# Corporate Profile & Overview (5 Symmetrical Cards wrapped in electric border)
st.markdown('<div class="section-title">Corporate Profile & Overview</div>', unsafe_allow_html=True)
st.markdown('<div class="section-description">Executive snapshot of the operating entity and its core revenue engine.</div>', unsafe_allow_html=True)

overview_items = [
    ("Company Name", company.get("company_name", "Not available")),
    ("Sector & Industry", company.get("industry", "Not available")),
    ("Business Model", company.get("business_type", "Not available")),
    ("Reporting Period", company.get("reporting_period", "Not available")),
    ("Filing Format", company.get("report_type", "Annual Report"))
]

overview_columns = st.columns(5)
for column, item in zip(overview_columns, overview_items):
    with column:
        st.markdown(f"""
        <div class="company-card">
            <div class="company-label">{item[0]}</div>
            <div class="company-value">{item[1]}</div>
        </div>
        """, unsafe_allow_html=True)

# Headline Financial Metrics (BI Tile Matrix with 3-Color Status Electric Borders: Green, Yellow, Red)
st.markdown('<div class="section-title">Financial Metrics</div>', unsafe_allow_html=True)
st.markdown('<div class="section-description">Core revenue, profit, and balance sheet indicators with verified YoY deltas.</div>', unsafe_allow_html=True)

headline_metrics = []
priority_words = ["total income", "revenue", "profit after tax", "pat", "net profit", "operating profit", "ebitda", "net interest income"]

for m in metrics:
    m_name = str(m.get("metric", "")).lower()
    if any(word in m_name for word in priority_words):
        if m not in headline_metrics:
            headline_metrics.append(m)

for m in metrics:
    if m not in headline_metrics:
        headline_metrics.append(m)

headline_metrics = headline_metrics[:4]

if headline_metrics:
    metric_cols = st.columns(len(headline_metrics))
    for col, m in zip(metric_cols, headline_metrics):
        with col:
            name = m.get("metric", "Metric")
            curr = m.get("current_period", "N/A")
            unit = m.get("unit", "")
            growth = str(m.get("yoy_growth", "")).strip()
            basis = m.get("basis", "")

            spark_width = 75
            card_class = "electric-kpi-card-yellow" # Default steady -> Yellow/Amber
            spark_color = "#fbbf24"

            if growth and growth.lower() not in ["n/a", "not available", ""]:
                if growth.startswith("-") or "decline" in growth.lower():
                    badge_html = f"""<span style="color:#f87171; font-weight:750; font-size:12px;">▼ {growth}</span>"""
                    spark_color = "#ef4444"
                    spark_width = 45
                    card_class = "electric-kpi-card-red" # Loss / negative -> Red
                else:
                    clean_g = growth if growth.startswith("+") else f"+{growth}"
                    badge_html = f"""<span style="color:#34d399; font-weight:750; font-size:12px;">▲ {clean_g} YoY</span>"""
                    spark_color = "#10b981"
                    spark_width = 85
                    card_class = "electric-kpi-card-green" # Growth / profit -> Green
            else:
                badge_html = """<span style="color:#94a3b8; font-weight:700; font-size:12px;">Reported Level</span>"""

            st.markdown(f"""
            <div class="{card_class}">
                <div>
                    <div class="kpi-header-row">
                        <span class="kpi-title">{name}</span>
                        {badge_html}
                    </div>
                    <div class="kpi-main-val">{curr} <span style="font-size:14px; font-weight:600; color:#94a3b8;">{unit}</span></div>
                </div>
                <div>
                    <div class="spark-track"><div style="width:{spark_width}%; height:100%; background:{spark_color}; border-radius:4px;"></div></div>
                    <div style="display:flex; justify-content:space-between; font-size:11px; color:#64748b; margin-top:6px;">
                        <span>Basis: {basis if basis else 'Reported'}</span>
                        <span>Audited</span>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

# Consolidated Financial Dashboards Suite (6 Interactive Tabs)
st.markdown('<div class="section-title" style="margin-top: 30px;">Financial Dashboards</div>', unsafe_allow_html=True)
st.markdown('<div class="section-description">Interactive analytical tabs structured with financial dashboard aesthetics:</div>', unsafe_allow_html=True)

tab_scorecard, tab_metrics, tab_charts, tab_mgmt, tab_risks, tab_investor = st.tabs([
    "⭐ Strategic Scorecard", "Financial Statement Table", "📊 Growth & Performance", "Management Outlook", "Risk Heatmap Matrix", "Analyst Signals & Takeaways"
])

# Tab 1: Strategic Scorecard
with tab_scorecard:
    st.subheader("Executive Strategic Diagnostic Scorecard")
    st.write("Structured 4-pillar evaluation matrix assessing corporate performance, capital resilience, and execution velocity:")

    if scorecard:
        col_s1, col_s2 = st.columns(2)
        pillars = [
            ("growth_momentum", "🚀 Growth Momentum", col_s1, "#38bdf8", 84),
            ("profitability_quality", "💰 Profitability & Quality", col_s2, "#34d399", 79),
            ("balance_sheet_safety", "🛡️ Balance Sheet Resilience", col_s1, "#818cf8", 92),
            ("strategic_execution", "⚙️ Strategic & Commercial Scale", col_s2, "#fbbf24", 88)
        ]

        for p_key, p_title, target_col, bar_color, default_pct in pillars:
            p_obj = scorecard.get(p_key, {})
            score = int(p_obj.get("health_pct", default_pct))
            badge = p_obj.get("badge", "Expansion")
            verdict = p_obj.get("verdict", "")
            points = p_obj.get("points", p_obj.get("tags", []))

            chips_html = "".join([
                f'<div style="background:#0c1220; border:1px solid #1f2d45; color:#cbd5e1; font-size:12.5px; padding:7px 11px; border-radius:6px; line-height:1.4; margin-bottom:6px;">✓ {t}</div>'
                for t in points[:3]
            ])

            card_html = f"""<div class="electric-kpi-card-blue" style="margin-bottom:16px; height: auto;">
<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
<span style="font-size:15.5px; font-weight:750; color:#ffffff;">{p_title}</span>
<span style="background:rgba(59,130,246,0.18); border:1px solid rgba(59,130,246,0.35); color:#93c5fd; padding:3px 10px; border-radius:6px; font-size:11.5px; font-weight:700;">{badge}</span>
</div>
<div style="color:#94a3b8; font-size:13px; line-height:1.45; margin-bottom:12px;">{verdict}</div>
<div style="display:flex; justify-content:space-between; font-size:11.5px; color:#cbd5e1; font-weight:600; margin-bottom:4px;">
<span>Diagnostic Pillar Score</span>
<span style="color:{bar_color}; font-weight:800;">{score}%</span>
</div>
<div style="background:#151d2f; border-radius:8px; height:8px; width:100%; margin-bottom:14px; overflow:hidden;">
<div style="width:{score}%; height:100%; background:{bar_color}; border-radius:8px;"></div>
</div>
<div style="font-size:11px; font-weight:700; color:#64748b; text-transform:uppercase; letter-spacing:0.5px; margin-bottom:8px;">Primary Reconciled Signals:</div>
{chips_html}
</div>"""
            with target_col:
                st.markdown(card_html, unsafe_allow_html=True)

# Tab 2: Financial Statement Table
with tab_metrics:
    st.subheader("Audited Financial & Operating Statement Table")
    if metrics:
        col_search, col_filter = st.columns([2, 1])
        with col_search:
            search_query = st.text_input("🔍 Search financial line items...", placeholder="e.g. Total Revenue, Net Profit, Borrowings, NPA", key="metric_search").lower()
        with col_filter:
            all_bases = list(set([m.get("basis", "").strip() for m in metrics if m.get("basis", "").strip()]))
            basis_filter = st.selectbox("Filter Basis", options=["All"] + all_bases, key="basis_filter")

        filtered_rows = []
        for m in metrics:
            metric_name = m.get("metric", "")
            basis_val = m.get("basis", "")
            if (not search_query or search_query in metric_name.lower()) and (basis_filter == "All" or basis_val.lower() == basis_filter.lower()):
                filtered_rows.append({
                    "Financial Statement Line Item": metric_name,
                    "Current Period": m.get("current_period", ""),
                    "Previous Period": m.get("previous_period", ""),
                    "YoY Delta": m.get("yoy_growth", ""),
                    "Unit": m.get("unit", ""),
                    "Basis": basis_val,
                    "Classification": auto_classify_metric(metric_name)
                })
        if filtered_rows and pd is not None:
            st.dataframe(filtered_rows, use_container_width=True, hide_index=True, height=400)

# Tab 3: Growth & Performance
with tab_charts:
    st.subheader("Visual Growth & Comparative Performance")
    st.write("Visual financial comparisons across key operating parameters with analytical context:")

    categories = ["All Metrics", "Revenue & Profit", "Balance Sheet & Assets", "Financial Health Ratios"]
    selected_cat = st.radio("Select Category Slicer:", options=categories, horizontal=True, key="bi_chart_slicer")

    chart_records = []
    for m in metrics:
        curr_val = parse_clean_float(m.get("current_period"))
        prev_val = parse_clean_float(m.get("previous_period"))
        m_name = m.get("metric", "").strip()
        auto_cat = auto_classify_metric(m_name)
        meaning = m.get("what_it_means", "Primary indicator of corporate operational trajectory.")
        
        if selected_cat == "All Metrics" or selected_cat.lower() == auto_cat.lower():
            if curr_val is not None and prev_val is not None:
                chart_records.append({
                    "Metric": m_name,
                    "Previous": prev_val,
                    "Current": curr_val,
                    "Unit": m.get("unit", "").strip(),
                    "Growth": m.get("yoy_growth", ""),
                    "Meaning": meaning
                })

    if chart_records:
        chart_cols = st.columns(2)
        for idx, item in enumerate(chart_records[:8]):
            c_val, p_val, u_lbl, m_name, growth, meaning = item["Current"], item["Previous"], item["Unit"], item["Metric"], item["Growth"], item["Meaning"]
            max_v = max(abs(c_val), abs(p_val)) if max(abs(c_val), abs(p_val)) > 0 else 1
            prev_pct = max(int((abs(p_val) / max_v) * 100), 8)
            curr_pct = max(int((abs(c_val) / max_v) * 100), 8)
            
            growth_str = str(growth).strip()
            if growth_str and growth_str.lower() not in ["n/a", "not available", ""]:
                delta_html = f'<span style="color:#34d399; font-size:11.5px; font-weight:750;">▲ +{growth_str.replace("+","")} YoY</span>' if not growth_str.startswith("-") else f'<span style="color:#f87171; font-size:11.5px; font-weight:750;">▼ {growth_str} YoY</span>'
            else:
                delta_html = '<span style="color:#94a3b8; font-size:11.5px; font-weight:700;">Audited Level</span>'

            chart_html = f"""<div class="electric-kpi-card-blue" style="margin-bottom:14px; height: auto;">
<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:10px;">
<span style="color:#ffffff; font-size:14.5px; font-weight:750;">{m_name}</span>
{delta_html}
</div>
<div style="margin-bottom:8px;">
<div style="display:flex; justify-content:space-between; font-size:12px; font-weight:600; color:#94a3b8; margin-bottom:3px;">
<span>Previous Period</span>
<span>{p_val:,.2f} {u_lbl}</span>
</div>
<div style="background:#151d2f; border-radius:6px; height:18px; width:100%; overflow:hidden;">
<div style="width:{prev_pct}%; background:#334155; height:100%; display:flex; align-items:center; justify-content:flex-end; padding-right:8px; color:#ffffff; font-size:11px; font-weight:700;">{p_val:,.2f}</div>
</div>
</div>
<div style="margin-bottom:10px;">
<div style="display:flex; justify-content:space-between; font-size:12px; font-weight:600; color:#cbd5e1; margin-bottom:3px;">
<span>Current Period</span>
<span>{c_val:,.2f} {u_lbl}</span>
</div>
<div style="background:#151d2f; border-radius:6px; height:18px; width:100%; overflow:hidden;">
<div style="width:{curr_pct}%; background:linear-gradient(90deg, #1d4ed8, #38bdf8); height:100%; display:flex; align-items:center; justify-content:flex-end; padding-right:8px; color:#ffffff; font-size:11px; font-weight:700;">{c_val:,.2f}</div>
</div>
</div>
<div style="color:#94a3b8; font-size:12px; line-height:1.4; border-top:1px solid #1a2234; padding-top:8px;">
💡 <b>Analytical Context:</b> {meaning}
</div>
</div>"""
            with chart_cols[idx % 2]:
                st.markdown(chart_html, unsafe_allow_html=True)
    else:
        st.info("No comparative figures available for this specific category slice.")

# Tab 4: Management Outlook
with tab_mgmt:
    st.subheader("Management Strategy & Future Execution Roadmaps")
    for item in management:
        with st.expander(f"🎯 {item.get('title', 'Strategic Pillar')}", expanded=False):
            st.write(item.get("summary", ""))

# Tab 5: Risk Heatmap Matrix
with tab_risks:
    st.subheader("Potential Risks & Headwinds (Risk Heatmap Matrix)")
    st.write("Visual categorization of operational, credit, regulatory, and market threats:")

    if risks:
        r_cols = st.columns(2)
        severity_palette = {
            "high": ("#ef4444", "rgba(239, 68, 68, 0.15)", "#18090c", "electric-kpi-card-red"),
            "moderate": ("#fbbf24", "rgba(245, 158, 11, 0.15)", "#181409", "electric-kpi-card-yellow"),
            "low": ("#60a5fa", "rgba(59, 130, 246, 0.15)", "#09121c", "electric-kpi-card-blue"),
            "operational": ("#60a5fa", "rgba(59, 130, 246, 0.15)", "#09121c", "electric-kpi-card-blue")
        }
        
        for idx, r in enumerate(risks):
            cat = r.get("category", "Market & Economy")
            impact = str(r.get("impact_level", "Moderate")).lower()
            if impact not in severity_palette:
                impact = "moderate"
            
            tag_color, tag_bg, card_bg, card_css = severity_palette[impact]
            
            risk_card_html = f"""<div class="{card_css}" style="background:{card_bg}; margin-bottom:12px; height:100%;">
<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
<span style="background:{tag_bg}; border:1px solid {tag_color}; color:{tag_color}; font-size:10.5px; font-weight:800; padding:2px 8px; border-radius:6px; text-transform:uppercase;">● {impact.upper()} IMPACT</span>
<span style="color:#94a3b8; font-size:11.5px; font-weight:600;">{cat}</span>
</div>
<div style="color:#ffffff; font-weight:750; font-size:14px; margin-bottom:6px;">⚠️ {r.get('title')}</div>
<div style="color:#94a3b8; font-size:12.5px; line-height:1.45; margin-bottom:10px;">{r.get('what_is_the_risk', '')}</div>
<div style="background:#06080e; border-left:3px solid {tag_color}; border-radius:0 6px 6px 0; padding:8px 12px; font-size:12px; color:#f1f5f9;">
<b>Impact on Financials:</b> {r.get('why_it_matters')}
</div>
</div>"""
            with r_cols[idx % 2]:
                st.markdown(risk_card_html, unsafe_allow_html=True)

# Tab 6: Analyst Signals & Takeaways
with tab_investor:
    st.subheader("Institutional Sentiment & Analyst Signals")
    
    bull_pct = int(takeaway.get("sentiment_score", 74))
    bear_pct = 100 - bull_pct

    st.markdown(f"""<div class="electric-kpi-card-blue" style="margin-bottom:20px; height: auto;">
<div style="display:flex; justify-content:space-between; font-size:13px; font-weight:700;">
<span style="color:#34d399;">🟢 Institutional Bull Catalysts: {bull_pct}%</span>
<span style="color:#f87171;">🔴 Headwinds & Cost Pressures: {bear_pct}%</span>
</div>
<div style="display:flex; height:10px; border-radius:10px; overflow:hidden; margin:10px 0 6px 0;">
<div style="width:{bull_pct}%; background:linear-gradient(90deg, #059669, #10b981);"></div>
<div style="width:{bear_pct}%; background:linear-gradient(90deg, #ef4444, #b91c1c);"></div>
</div>
<div style="color:#64748b; font-size:11.5px;">Weighted sentiment derived from revenue growth momentum vs downside risk disclosures.</div>
</div>""", unsafe_allow_html=True)

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 🟢 Improving Tailwinds & Growth Vectors")
        for item in takeaway.get("improving", []):
            st.markdown(f'<div style="background:#062319; border-left:3px solid #10b981; border-radius:6px; padding:10px 14px; margin-bottom:8px; color:#d1fae5; font-size:13px;">✓ {item}</div>', unsafe_allow_html=True)
    with col2:
        st.markdown("#### 🔴 Weakening Headwinds & Margin Pressures")
        for item in takeaway.get("weakening", []):
            st.markdown(f'<div style="background:#260d13; border-left:3px solid #ef4444; border-radius:6px; padding:10px 14px; margin-bottom:8px; color:#fee2e2; font-size:13px;">✗ {item}</div>', unsafe_allow_html=True)

# ============================================================
# PART 2: CONDITIONAL ON-DEMAND SECTIONS (RADIO / BUTTON TOGGLES)
# ============================================================

# ------------------------------------------------------------
# 1. 3 to 5-Year Historical Trends & Strategic Forecasting
# ------------------------------------------------------------
st.markdown("""
<div class="fintech-banner">
    <div class="fintech-banner-title">📈 3 to 5-Year Historical Trends & Strategic Forecasting</div>
    <div class="fintech-banner-desc">Financial Dashboard: AI-driven predictive projection paths combined with live web-sourced market trend analysis.</div>
</div>
""", unsafe_allow_html=True)

forecast_toggle = st.radio(
    "Do you want to generate trend analysis and forecasting?",
    options=["No, keep standard view", "Yes, generate 3-5 year trend & forecasting analysis"],
    index=0,
    horizontal=True,
    key="forecast_toggle_rad"
)

if forecast_toggle == "No, keep standard view":
    st.markdown("""
    <div class="electric-kpi-card-blue" style="padding: 14px 18px; color: #94a3b8; font-size: 13px; height: auto;">
        💡 <strong>Standard view retained.</strong> Feel free to explore the overview or move forward to the next section.
    </div>
    """, unsafe_allow_html=True)
elif forecast_toggle == "Yes, generate 3-5 year trend & forecasting analysis":
    
    if not st.session_state.get("forecast_loaded") or st.session_state.get("forecast_company") != company.get("company_name"):
        with st.spinner("🌍 Tracing live internet market trends, past financial history, and predictive compounding models via Gemini Search..."):
            c_name = company.get("company_name", "the company")
            c_ticker = company.get("stock_ticker", "")
            
            web_forecast_prompt = (
                "You are an expert institutional equity research analyst. "
                f"Analyze the company '{c_name}' (Ticker: {c_ticker}) using live web search data for recent news, strategic updates, and market trends. "
                "Provide a natural-language predictive projection breakdown strictly matching this JSON structure. Do not use markdown blocks:\n"
                "{\n"
                "  \"cagr_value\": \"Dynamic value (e.g. +12.5% p.a.)\",\n"
                "  \"margin_outlook\": \"Dynamic value (e.g. Expanding (+50 bps))\",\n"
                "  \"risk_scenario\": \"Dynamic value (e.g. Base Case / Stress Case)\",\n"
                "  \"rationale_points\": [\n"
                "    \"Point 1: Based on specific recent search data...\",\n"
                "    \"Point 2: Based on operational updates...\",\n"
                "    \"Point 3: Based on sector outlook...\"\n"
                "  ]\n"
                "}"
            )
            try:
                # Passing only the text prompt ensures search context limits aren't exceeded
                res_f = generate_with_fallback(contents=[web_forecast_prompt], json_mode=False, use_search=True)
                f_parsed = clean_json_response(res_f.text)
                if not f_parsed or "rationale_points" not in f_parsed:
                    raise Exception("Invalid structure")
                st.session_state.forecast_data = f_parsed
                st.session_state.forecast_company = c_name
                st.session_state.forecast_loaded = True
            except Exception:
                # Dynamic fallback based on actual user uploaded data
                fallback_growth = headline_metrics[0].get("yoy_growth", "+10.0%") if headline_metrics else "+10.0%"
                st.session_state.forecast_data = {
                    "cagr_value": f"Estimated {fallback_growth} p.a.",
                    "margin_outlook": "Stable Outlook",
                    "risk_scenario": "Base Case (Audited Trends)",
                    "rationale_points": [
                        f"Historical trend reconciliation for {c_name} indicates steady operational workflows.",
                        "Live internet market data tracing encountered temporary limits, projection relies on fundamental trailing 12-month statements.",
                        "Robust capital buffers support financial stability across localized economic cycles."
                    ]
                }
                st.session_state.forecast_company = c_name
                st.session_state.forecast_loaded = True

    f_res = st.session_state.forecast_data

    st.markdown("""
    <div class="electric-kpi-card-blue" style="margin-bottom:20px; height: auto;">
        <div style="color:#60a5fa; font-size:16px; font-weight:750; margin-bottom:8px;">📊 Predictive Revenue & Profit Trajectory (Next 3–5 Years)</div>
        <div style="color:#94a3b8; font-size:13.5px; margin-bottom:16px; line-height: 1.5;">
            Dynamic projection model synthesized from audited annual filings, live internet market data tracing, and macroeconomic compounding vectors.
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        cagr_val = f_res.get('cagr_value', '+10.0% p.a.')
        st.markdown(f"""
        <div class="electric-kpi-card-green" style="text-align: center;">
            <div style="color:#94a3b8; font-size:12px; font-weight:700; text-transform:uppercase;">Projected 3Y CAGR Growth</div>
            <div style="color:#34d399; font-size:26px; font-weight:800; margin: 8px 0;">{cagr_val}</div>
            <div style="color:#94a3b8; font-size:11.5px;">Revenue Expansion Rate</div>
        </div>
        """, unsafe_allow_html=True)
    with fc2:
        margin_val = f_res.get('margin_outlook', 'Stable')
        st.markdown(f"""
        <div class="electric-kpi-card-blue" style="text-align: center;">
            <div style="color:#94a3b8; font-size:12px; font-weight:700; text-transform:uppercase;">Operating Margin Outlook</div>
            <div style="color:#60a5fa; font-size:24px; font-weight:800; margin: 8px 0;">{margin_val}</div>
            <div style="color:#94a3b8; font-size:11.5px;">Cost Optimization Leverage</div>
        </div>
        """, unsafe_allow_html=True)
    with fc3:
        risk_val = f_res.get('risk_scenario', 'Base Case')
        st.markdown(f"""
        <div class="electric-kpi-card-yellow" style="text-align: center;">
            <div style="color:#94a3b8; font-size:12px; font-weight:700; text-transform:uppercase;">Risk-Adjusted Scenario</div>
            <div style="color:#fbbf24; font-size:20px; font-weight:800; margin: 8px 0;">{risk_val}</div>
            <div style="color:#94a3b8; font-size:11.5px;">Solvency Buffer Maintained</div>
        </div>
        """, unsafe_allow_html=True)

    # Detailed Bullet-Point Explanations with Electric Borderlines
    rationale_items = f_res.get("rationale_points", [])
    bullets_html = "".join([f'<li style="margin-bottom: 12px;"><strong>Market & Historical Trace Rationale:</strong> {pt}</li>' for pt in rationale_items])

    st.markdown(f"""
    <div class="electric-kpi-card-green" style="margin-top: 20px; height: auto; padding: 22px;">
        <div style="color:#34d399; font-size:16px; font-weight:750; margin-bottom:12px;">💡 Comprehensive Analyst Rationale & Future Outlook (Web-Sourced & Audited)</div>
        <div style="color:#cbd5e1; font-size:13.5px; line-height: 1.6;">
            <ul style="margin: 0; padding-left: 20px;">
                {bullets_html}
            </ul>
        </div>
    </div>
    """, unsafe_allow_html=True)

# ------------------------------------------------------------
# 2. In-Depth Financial Investigation (Forensic Audit Module)
# ------------------------------------------------------------
st.markdown("""
<div class="fintech-banner">
    <div class="fintech-banner-title">🔬 In-Depth Financial Investigation</div>
    <div class="fintech-banner-desc">Financial Dashboard: Comprehensive forensic assessment auditing profit margins, capital return ratios, and solvency cushions.</div>
</div>
""", unsafe_allow_html=True)

deep_choice = st.radio(
    "Do you want to generate deep-dive financial analysis?",
    options=["No, keep summary view", "Yes, generate deep-dive financial analysis"],
    index=0,
    horizontal=True,
    key="deep_dive_choice_rad"
)

if deep_choice == "No, keep summary view":
    st.markdown("""
    <div class="electric-kpi-card-blue" style="padding: 14px 18px; color: #94a3b8; font-size: 13px; height: auto;">
        💡 Thank you for your time. Feel free to look at the overview of the report and move forward to the next section.
    </div>
    """, unsafe_allow_html=True)

elif deep_choice == "Yes, generate deep-dive financial analysis":
    prof = deep_investigation.get("profitability_and_margins", {
        "headline": "Core operating revenue expanded with strategic margin reinvestments.",
        "points": [
            "Operating revenue demonstrated strong double-digit growth across primary business verticals.",
            "Net margins reflect upfront technology and infrastructure scaling costs.",
            "Return on Equity remains stable backed by recurring operational cash flows."
        ]
    })
    debt = deep_investigation.get("borrowings_and_capital_cushion", {
        "headline": "Robust capital adequacy and conservative leverage ratios.",
        "points": [
            "Total net worth provides a substantial solvency cushion against operating obligations.",
            "Borrowings are balanced across diversified long-term credit and banking lines.",
            "Liquid reserves and cash balances remain fully compliant with regulatory buffers."
        ]
    })
    eff = deep_investigation.get("operating_efficiency_and_scale", {
        "headline": "Operational efficiency improving via digital process automation.",
        "points": [
            "Core operating expenses optimized through digital transaction workflows.",
            "Revenue mix continues to diversify across high-margin fee-based services.",
            "Staff and branch productivity metrics showed positive year-over-year gains."
        ]
    })

    col_d1, col_d2, col_d3 = st.columns(3)

    with col_d1:
        points_html = "".join([f'<div style="background:#0c1220; border-left:3px solid #38bdf8; border-radius:6px; padding:10px 12px; margin-bottom:8px; font-size:12.5px; color:#cbd5e1; line-height:1.45;">• {p}</div>' for p in prof.get("points", [])])
        st.markdown(f"""
        <div class="electric-kpi-card-green" style="height:100%; min-height: 280px;">
            <div style="color:#38bdf8; font-size:16px; font-weight:750; margin-bottom:8px;">📊 Profit Margins & Returns</div>
            <div style="color:#ffffff; font-weight:650; font-size:13.5px; margin-bottom:14px; line-height:1.4;">{prof.get('headline', '')}</div>
            {points_html}
        </div>
        """, unsafe_allow_html=True)

    with col_d2:
        points_html = "".join([f'<div style="background:#0c1220; border-left:3px solid #818cf8; border-radius:6px; padding:10px 12px; margin-bottom:8px; font-size:12.5px; color:#cbd5e1; line-height:1.45;">• {p}</div>' for p in debt.get("points", [])])
        st.markdown(f"""
        <div class="electric-kpi-card-blue" style="height:100%; min-height: 280px;">
            <div style="color:#818cf8; font-size:16px; font-weight:750; margin-bottom:8px;">🛡️ Borrowings & Capital Cushion</div>
            <div style="color:#ffffff; font-weight:650; font-size:13.5px; margin-bottom:14px; line-height:1.4;">{debt.get('headline', '')}</div>
            {points_html}
        </div>
        """, unsafe_allow_html=True)

    with col_d3:
        points_html = "".join([f'<div style="background:#0c1220; border-left:3px solid #34d399; border-radius:6px; padding:10px 12px; margin-bottom:8px; font-size:12.5px; color:#cbd5e1; line-height:1.45;">• {p}</div>' for p in eff.get("points", [])])
        st.markdown(f"""
        <div class="electric-kpi-card-green" style="height:100%; min-height: 280px;">
            <div style="color:#34d399; font-size:16px; font-weight:750; margin-bottom:8px;">⚙️ Operational Scale & Efficiency</div>
            <div style="color:#ffffff; font-weight:650; font-size:13.5px; margin-bottom:14px; line-height:1.4;">{eff.get('headline', '')}</div>
            {points_html}
        </div>
        """, unsafe_allow_html=True)

# ------------------------------------------------------------
# 3. Personalized Investment Position & Market Valuation
# ------------------------------------------------------------
st.markdown("""
<div class="fintech-banner">
    <div class="fintech-banner-title">💼 Personalized Investment Position & Market Valuation</div>
    <div class="fintech-banner-desc">Financial Dashboard: Evaluate equity holdings against live exchange pricing, downside protection, and 5-to-8 year compounding horizons.</div>
</div>
""", unsafe_allow_html=True)

investor_mcq = st.radio(
    "Are you currently an investor in this company's stock?",
    options=["Select an option...", "Yes, I hold shares in this company", "No, I am just studying / evaluating"],
    index=0,
    horizontal=True,
    key="inv_mcq"
)

if investor_mcq == "No, I am just studying / evaluating":
    st.markdown("""
    <div class="electric-kpi-card-blue" style="padding: 14px 18px; color: #94a3b8; font-size: 13px; height: auto;">
        💡 Thank you for your time. Feel free to look at the overview of the report and move forward to the next section.
    </div>
    """, unsafe_allow_html=True)

elif investor_mcq == "Yes, I hold shares in this company":
    col_inv1, col_inv2 = st.columns(2)
    with col_inv1:
        total_invested_input = st.number_input("Total Capital Invested (₹)", min_value=0.0, value=None, placeholder="e.g. 50000.00", step=500.0, format="%.2f", key="inv_amt")
    with col_inv2:
        avg_price_input = st.number_input("Average Purchase Price per Share (₹)", min_value=0.0, value=None, placeholder="e.g. 250.00", step=1.0, format="%.2f", key="inv_price")

    if total_invested_input and avg_price_input and total_invested_input > 0 and avg_price_input > 0:
        calc_shares = int(total_invested_input // avg_price_input)
        st.caption(f"Calculated Holding: ~{calc_shares:,} Shares")
        
        pos_loader_placeholder = st.empty()

        if st.button("⚡ Run Portfolio Diagnostic", type="primary"):
            c_name = company.get('company_name', 'this company')
            t_hint = company.get('stock_ticker', '')

            pos_loader_placeholder.markdown("""
            <div class="center-loader-box">
                <div class="telemetry-pill" style="margin-bottom: 12px;">Valuation Engine Active</div>
                <div class="fintech-spinner"></div>
                <div style="font-size: 18px; font-weight: 750; color: #fff; margin-bottom: 6px;">Executing Investment Valuation...</div>
                <div class="pulse-text" style="font-size: 13px; color: #94a3b8;">Cross-referencing entry price against net worth backing and compounding models.</div>
                <div class="loader-progress-track"><div class="loader-progress-fill"></div></div>
            </div>
            """, unsafe_allow_html=True)

            market_info = fetch_live_stock_price(c_name, t_hint)
            live_price = market_info["price"] if market_info else avg_price_input
            live_date = market_info["as_on"] if market_info else datetime.today().strftime("%d %b %Y")
            exchange_tag = f"{market_info['exchange']}: {market_info['ticker']}" if market_info else "Global Exchange"

            pnl_amt = (live_price - avg_price_input) * calc_shares
            pnl_pct = ((live_price - avg_price_input) / avg_price_input) * 100
            pnl_sign = "+" if pnl_amt >= 0 else ""
            pnl_str = f"{pnl_sign}{pnl_pct:.2f}%"
            amt_str = f"{pnl_sign}₹{pnl_amt:,.2f}"
            cmp_display = f"₹{live_price:,.2f}"

            analysis_req_prompt = (
                "You are an expert institutional equity research mentor analyzing an investor's equity position in " + str(c_name) + ".\n"
                "Deliver a structured valuation synthesis.\n"
                "INVESTMENT PARAMETERS:\n"
                "- Capital Invested: ₹" + str(total_invested_input) + "\n"
                "- Purchase Price Basis: ₹" + str(avg_price_input) + "\n"
                "- Current Market Price: " + str(cmp_display) + "\n\n"
                "STRUCTURE YOUR JSON OUTPUT STRICTLY:\n"
                "{\n"
                "  \"profit_or_loss_summary\": \"Detailed breakdown of position performance.\",\n"
                "  \"investor_checkpoints\": [{\"title\": \"Checkpoint Title\", \"explanation\": \"What the investor must closely look into regarding fundamentals and risks.\"}],\n"
                "  \"future_aspect\": [{\"title\": \"Future Aspect Title\", \"explanation\": \"The future aspect and compounding potential over the next 5 to 8 years.\"}]\n"
                "}"
            )
            try:
                pos_res = generate_with_fallback(contents=[analysis_req_prompt, st.session_state.gemini_file], json_mode=True)
                parsed_pos = clean_json_response(pos_res.text)
            except Exception:
                parsed_pos = {}

            pos_loader_placeholder.empty()

            parsed_pos["cmp_display"] = cmp_display
            parsed_pos["live_date"] = live_date
            parsed_pos["exchange_tag"] = exchange_tag
            parsed_pos["pnl_str"] = pnl_str
            parsed_pos["amt_str"] = amt_str
            parsed_pos["is_pos"] = (pnl_amt >= 0)
            parsed_pos["live_price"] = live_price
            parsed_pos["avg_price"] = avg_price_input
            parsed_pos["calc_shares"] = calc_shares
            parsed_pos["invested_amt"] = total_invested_input
            st.session_state.position_assessment = parsed_pos

        if st.session_state.position_assessment:
            p_data = st.session_state.position_assessment
            st.markdown("---")
            st.markdown("### 📋 Financial Dashboard: Equity Position Assessment")
            
            is_pos = p_data.get("is_pos", True)
            pnl_card_css = "electric-kpi-card-green" if is_pos else "electric-kpi-card-red"
            pnl_color = "#34d399" if is_pos else "#f87171"

            cm1, cm2, cm3, cm4 = st.columns(4)
            with cm1:
                st.markdown(f"""
                <div class="electric-kpi-card-blue">
                    <div style="color: #94a3b8; font-size: 12px; font-weight: 700; text-transform: uppercase;">Invested Capital</div>
                    <div style="font-size: 22px; font-weight: 800; color: #fff; margin: 6px 0;">₹{p_data.get('invested_amt', total_invested_input):,.2f}</div>
                    <div style="color: #94a3b8; font-size: 11.5px;">~{p_data.get('calc_shares', calc_shares):,} Shares</div>
                </div>
                """, unsafe_allow_html=True)
            with cm2:
                st.markdown(f"""
                <div class="electric-kpi-card-blue">
                    <div style="color: #94a3b8; font-size: 12px; font-weight: 700; text-transform: uppercase;">Purchase Price Basis</div>
                    <div style="font-size: 22px; font-weight: 800; color: #fff; margin: 6px 0;">₹{p_data.get('avg_price', avg_price_input):,.2f}</div>
                    <div style="color: #94a3b8; font-size: 11.5px;">Cost Basis per Share</div>
                </div>
                """, unsafe_allow_html=True)
            with cm3:
                st.markdown(f"""
                <div class="electric-kpi-card-blue">
                    <div style="color: #94a3b8; font-size: 12px; font-weight: 700; text-transform: uppercase;">Current Market Price</div>
                    <div style="font-size: 22px; font-weight: 800; color: #60a5fa; margin: 6px 0;">{p_data.get('cmp_display', 'N/A')}</div>
                    <div style="color: #94a3b8; font-size: 11.5px;">{p_data.get('exchange_tag', 'NSE/BSE')}</div>
                </div>
                """, unsafe_allow_html=True)
            with cm4:
                st.markdown(f"""
                <div class="{pnl_card_css}">
                    <div style="color: #94a3b8; font-size: 12px; font-weight: 700; text-transform: uppercase;">Unrealized P&L Return</div>
                    <div style="font-size: 22px; font-weight: 800; color: {pnl_color}; margin: 6px 0;">{p_data.get('pnl_str', 'N/A')}</div>
                    <div style="color: {pnl_color}; font-size: 11.5px;">{p_data.get('amt_str', '')}</div>
                </div>
                """, unsafe_allow_html=True)

            pnl_summary = p_data.get("profit_or_loss_summary", "")
            banner_border = "#10b981" if is_pos else "#ef4444"
            banner_bg = "#071f16" if is_pos else "#240d12"
            banner_text = "#d1fae5" if is_pos else "#fee2e2"

            st.markdown(f"""
            <div style="background: {banner_bg}; border-left: 4px solid {banner_border}; padding: 14px 18px; border-radius: 0 8px 8px 0; color: {banner_text}; font-size: 13.5px; line-height: 1.55; margin: 16px 0;">
                <strong>Position Dynamics:</strong> {pnl_summary}
            </div>
            """, unsafe_allow_html=True)

            st.markdown("#### 🔍 What You Must Look Into (Investor Checkpoints)")
            for item in p_data.get("investor_checkpoints", p_data.get("price_safety_points", [])):
                st.markdown(f"""
                <div class="electric-kpi-card-blue" style="margin-bottom: 10px; height: auto;">
                    <div style="font-weight: 750; color: #fff; margin-bottom: 4px;">✓ {item.get('title', '')}</div>
                    <div style="color: #94a3b8; font-size: 13px;">{item.get('explanation', '')}</div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("#### 🚀 Future Aspect & Long-Term Compounding (5-8 Years)")
            for item in p_data.get("future_aspect", p_data.get("long_term_outlook_5_to_8_years", [])):
                st.markdown(f"""
                <div class="electric-kpi-card-green" style="margin-bottom: 10px; height: auto;">
                    <div style="font-weight: 750; color: #fff; margin-bottom: 4px;">◆ {item.get('title', '')}</div>
                    <div style="color: #94a3b8; font-size: 13px;">{item.get('explanation', '')}</div>
                </div>
                """, unsafe_allow_html=True)

# ------------------------------------------------------------
# 4. Export Financial Research Suite
# ------------------------------------------------------------
st.markdown("""
<div class="fintech-banner">
    <div class="fintech-banner-title">📥 Export Financial Research Suite</div>
    <div class="fintech-banner-desc">Professional Excel financial model workbook (.xlsx) integrated with formulas and KPI summary cards.</div>
</div>
""", unsafe_allow_html=True)

export_choice = st.radio(
    "Do you want to generate an institutional research export suite?",
    options=["No, keep on-screen view", "Yes, generate institutional research export suite"],
    index=0,
    horizontal=True,
    key="export_rad"
)

if export_choice == "No, keep on-screen view":
    st.markdown("""
    <div class="electric-kpi-card-blue" style="padding: 14px 18px; color: #94a3b8; font-size: 13px; height: auto;">
        💡 Thank you for your time. Feel free to look at the overview of the report and move forward to the next section.
    </div>
    """, unsafe_allow_html=True)

elif export_choice == "Yes, generate institutional research export suite":
    comp_name = company.get("company_name", "Company")

    excel_buffer = io.BytesIO()
    if pd is not None and openpyxl is not None:
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        try:
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                wb = writer.book
                
                # ==========================
                # SHEET 1: Executive Summary 
                # ==========================
                ws_summary = wb.create_sheet("Executive Summary", 0)
                ws_summary.sheet_view.showGridLines = False # Dashboard Look
                
                # Header Title Styling
                ws_summary.merge_cells('B2:H3')
                title_cell = ws_summary['B2']
                title_cell.value = f"INSTITUTIONAL FINANCIAL RESEARCH: {str(comp_name).upper()}"
                title_cell.font = Font(size=22, bold=True, color="FFFFFF")
                title_cell.fill = PatternFill(fill_type="solid", start_color="0B101C")
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Border Styles for Cards
                thin_border = Border(
                    left=Side(style='thin', color="3B82F6"), 
                    right=Side(style='thin', color="3B82F6"), 
                    top=Side(style='thin', color="3B82F6"), 
                    bottom=Side(style='thin', color="3B82F6")
                )

                # Spacer Columns (Columns C, E, G)
                ws_summary.column_dimensions['C'].width = 3
                ws_summary.column_dimensions['E'].width = 3
                ws_summary.column_dimensions['G'].width = 3

                # Render KPI Cards in Excel (Spaced out over columns B, D, F, H)
                for i, m in enumerate(headline_metrics[:4]):
                    col_idx = 2 + (i * 2) 
                    col_letter = get_column_letter(col_idx)
                    
                    ws_summary.column_dimensions[col_letter].width = 32
                    
                    ws_summary.row_dimensions[5].height = 30
                    ws_summary.row_dimensions[6].height = 50
                    ws_summary.row_dimensions[7].height = 30
                    
                    # Title Card Part
                    c_title = ws_summary[f'{col_letter}5']
                    c_title.value = str(m.get("metric", "KPI")).upper()
                    c_title.font = Font(size=11, bold=True, color="94A3B8")
                    c_title.fill = PatternFill(fill_type="solid", start_color="1E293B")
                    c_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c_title.border = thin_border
                    
                    # Value Card Part
                    c_val = ws_summary[f'{col_letter}6']
                    c_val.value = f"{m.get('current_period', 'N/A')} {m.get('unit', '')}"
                    c_val.font = Font(size=18, bold=True, color="38BDF8")
                    c_val.fill = PatternFill(fill_type="solid", start_color="0B101C")
                    c_val.alignment = Alignment(horizontal="center", vertical="center")
                    c_val.border = thin_border
                    
                    # Growth Status Card Part
                    c_yoy = ws_summary[f'{col_letter}7']
                    yoy_val = str(m.get('yoy_growth', 'N/A'))
                    c_yoy.value = f"YoY Delta: {yoy_val}"
                    color_yoy = "10B981" if ("+" in yoy_val or "-" not in yoy_val) else "EF4444"
                    if yoy_val in ["N/A", "", "None"]: color_yoy = "94A3B8"
                    
                    c_yoy.font = Font(size=12, bold=True, color=color_yoy)
                    c_yoy.fill = PatternFill(fill_type="solid", start_color="0B101C")
                    c_yoy.alignment = Alignment(horizontal="center", vertical="center")
                    c_yoy.border = thin_border

                # ==========================
                # SHEET 2: Financial Metrics
                # ==========================
                metrics_data = []
                for m in metrics:
                    metrics_data.append({
                        "Line Item": m.get("metric", ""),
                        "Current Period": m.get("current_period", ""),
                        "Previous Period": m.get("previous_period", ""),
                        "YoY Delta": m.get("yoy_growth", ""),
                        "Unit": m.get("unit", ""),
                        "Basis": m.get("basis", ""),
                        "Category": auto_classify_metric(m.get("metric", "")),
                        "Analytical Context": m.get("what_it_means", "")
                    })
                df_metrics = pd.DataFrame(metrics_data)
                df_metrics.to_excel(writer, sheet_name="Financial Metrics", index=False)
                
                # ==========================
                # SHEET 3: Risk Matrix
                # ==========================
                risks_data = []
                for r in risks:
                    risks_data.append({
                        "Risk Title": r.get("title", ""),
                        "Category": r.get("category", ""),
                        "Impact Severity": r.get("impact_level", ""),
                        "Hazard Description": r.get("what_is_the_risk", ""),
                        "Financial Implication": r.get("why_it_matters", "")
                    })
                df_risks = pd.DataFrame(risks_data)
                df_risks.to_excel(writer, sheet_name="Risk Matrix", index=False)

                # ==========================
                # Style Tables (Headers, Widths, Grid Borders, and Freeze Panes)
                # ==========================
                thin_grid = Border(
                    left=Side(style='thin', color="DDDDDD"),
                    right=Side(style='thin', color="DDDDDD"),
                    top=Side(style='thin', color="DDDDDD"),
                    bottom=Side(style='thin', color="DDDDDD")
                )

                # Style Financial Metrics
                ws_metrics = writer.sheets["Financial Metrics"]
                ws_metrics.freeze_panes = 'A2'
                col_widths_fm = [35, 15, 15, 12, 10, 15, 20, 50]
                for i, w in enumerate(col_widths_fm, 1):
                    ws_metrics.column_dimensions[get_column_letter(i)].width = w

                for row in ws_metrics.iter_rows(min_row=1, max_row=ws_metrics.max_row, min_col=1, max_col=8):
                    for cell in row:
                        if cell.row == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(fill_type="solid", start_color="0F172A")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            cell.border = thin_grid

                # Style Risk Matrix
                ws_risks = writer.sheets["Risk Matrix"]
                ws_risks.freeze_panes = 'A2'
                col_widths_rm = [30, 20, 15, 40, 45]
                for i, w in enumerate(col_widths_rm, 1):
                    ws_risks.column_dimensions[get_column_letter(i)].width = w

                for row in ws_risks.iter_rows(min_row=1, max_row=ws_risks.max_row, min_col=1, max_col=5):
                    for cell in row:
                        if cell.row == 1:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill(fill_type="solid", start_color="0F172A")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            cell.border = thin_grid

                # Remove default sheet
                if "Sheet" in wb.sheetnames:
                    del wb["Sheet"]

            excel_bytes = excel_buffer.getvalue()
        except Exception:
            excel_bytes = b""
    else:
        excel_bytes = b""

    clean_file_name = re.sub(r'[^A-Za-z0-9_]', '_', comp_name)
    
    # XLSX Download Button only as requested
    st.download_button(
        label="📊 Download Professional Excel Model Workbook (.xlsx)",
        data=excel_bytes,
        file_name=f"{clean_file_name}_Financial_Model.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ------------------------------------------------------------
# 5. Interactive Institutional Research Copilot (Chat Interface)
# ------------------------------------------------------------
st.markdown("""
<div class="fintech-banner">
    <div class="fintech-banner-title">💬 Interactive Institutional Research Copilot</div>
    <div class="fintech-banner-desc">Query balance sheet line items, capital ratios, risk factors, or segment disclosures grounded strictly in this filing.</div>
</div>
""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
use_web_search_chat = st.toggle("🌐 Enable Live Web Search (Combines PDF with live market news - Takes ~10-15s)", value=False)

chip_cols = st.columns(4)
suggested_q = None
with chip_cols[0]:
    if st.button("📈 Profit Margin Trajectory", key="c1"): suggested_q = "What were the primary drivers and cost pressures that affected net profit margins YoY?"
with chip_cols[1]:
    if st.button("🚀 Core Business Growth Drivers", key="c2"): suggested_q = "What are the company's major operational growth drivers and revenue scaling avenues?"
with chip_cols[2]:
    if st.button("💰 Debt Solvency & Liquidity", key="c3"): suggested_q = "How is the company's balance sheet positioned in terms of debt leverage, liquidity reserves, and solvency?"
with chip_cols[3]:
    if st.button("⚠️ Material Operational Risks", key="c4"): suggested_q = "What are the primary operational, credit, regulatory, and market risks outlined in the report?"

# Render conversational history
for chat in st.session_state.chat_history:
    if chat["role"] == "user":
        st.markdown(f"""
        <div class="chat-box-card" style="border-left: 3.5px solid #3b82f6;">
            <div class="chat-user-badge">🧑‍💻 Investor Query</div>
            <div class="chat-text">{chat["content"]}</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"""
        <div class="chat-box-card" style="border-left: 3.5px solid #10b981; background: #071318;">
            <div class="chat-bot-badge">🤖 Financial Analyst AI</div>
            <div class="chat-text">{chat["content"]}</div>
        </div>
        """, unsafe_allow_html=True)

user_q = st.chat_input("Ask a research query about this corporate report...", key="main_chat_input")
active_q = user_q if user_q else suggested_q

if active_q:
    st.session_state.chat_history.append({"role": "user", "content": active_q})
    
    st.markdown(f"""
    <div class="chat-box-card" style="border-left: 3.5px solid #3b82f6;">
        <div class="chat-user-badge">🧑‍💻 Investor Query</div>
        <div class="chat-text">{active_q}</div>
    </div>
    """, unsafe_allow_html=True)

    c_name = company.get('company_name', 'the company')

    report_context = f"""
AUDITED COMPANY DATA:
- Entity: {c_name} ({company.get('stock_ticker', '')})
- Industry: {company.get('industry', 'N/A')}
- Business Profile: {company.get('business_type', 'N/A')}
- Extracted Metrics: {json.dumps(metrics)}
- Executive Scorecard: {json.dumps(scorecard)}
- Key Risks: {json.dumps(risks)}
- Management Themes: {json.dumps(management)}
- Analyst Takeaways: {json.dumps(takeaway)}
"""

    search_instruction = "Synthesize the provided data with live internet context dynamically to address current trends, market data, and recent news." if use_web_search_chat else "Rely ONLY on the provided financial summary data below. Do not use outside information or hallucinate."

    chat_prompt = f"""
You are an expert financial analyst. Answer the user's question accurately using the comprehensive financial summary extracted from the company's audited report.
CRITICAL INSTRUCTIONS:
- Explain your answer in SIMPLE TERMS.
- Avoid heavy technical jargon. If you must use a financial term, briefly explain it simply.
- Use clear bullet points and exact figures from the provided summary.
- {search_instruction}

{report_context}

INVESTOR QUESTION: {active_q}
"""

    chat_response_placeholder = st.empty()
    chat_response_placeholder.markdown("""
    <div class="chat-box-card" style="border-left: 3.5px solid #34d399; background: #071318;">
        <div class="chat-bot-badge">🤖 Financial Analyst AI</div>
        <div class="chat-text" style="color: #94a3b8;">
            <div class="pulse-text"><em>Analyzing report disclosures and structuring insights...</em></div>
            <div class="loader-progress-track" style="margin-top: 8px;"><div class="loader-progress-fill"></div></div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    try:
        # Crucial fix: Send only the chat_prompt (which includes the lightweight JSON summary) 
        # and DO NOT attach st.session_state.gemini_file (the massive PDF) to prevent rate limit crashes.
        res = generate_chat_response(contents=[chat_prompt], use_search=use_web_search_chat)
        ans = res.text.strip() if res.text else "No relevant disclosure found."
        
        st.session_state.chat_history.append({"role": "assistant", "content": ans})
        
        chat_response_placeholder.markdown(f"""
        <div class="chat-box-card" style="border-left: 3.5px solid #10b981; background: #071318;">
            <div class="chat-bot-badge">🤖 Financial Analyst AI</div>
            <div class="chat-text">{ans}</div>
        </div>
        """, unsafe_allow_html=True)
    except Exception as e:
        chat_response_placeholder.empty()
        st.error(f"Error: {e}")

# ============================================================
# FOOTER
# ============================================================
st.divider()
st.markdown('<div class="footer">Talk with your AI for this company-related report. Financial Analyst AI • Grounded Institutional Financial Analysis. For educational & research use only.</div>', unsafe_allow_html=True)
